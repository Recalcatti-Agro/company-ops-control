from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import re
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from core.models import (
    ArsUsdConversion,
    Commitment,
    Currency,
    Expense,
    Investment,
    Investor,
    Reinvestment,
    Work,
    WorkParticipation,
)


def d(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except Exception:
        return None


def dt(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def s(value: Any) -> str:
    return str(value).strip() if value is not None else ""


class Command(BaseCommand):
    help = "Importa datos iniciales desde Excel histórico"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default="/app/data/R - numeros.xlsx",
            help="Ruta al archivo xlsx (dentro del contenedor)",
        )
        parser.add_argument(
            "--no-reset",
            action="store_true",
            help="No borrar datos existentes antes de importar",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = Path(options["file"])
        if not file_path.exists():
            raise CommandError(f"No existe el archivo: {file_path}")

        wb = load_workbook(file_path, data_only=True)

        if not options["no_reset"]:
            WorkParticipation.objects.all().delete()
            Reinvestment.objects.all().delete()
            Work.objects.all().delete()
            Commitment.objects.all().delete()
            Expense.objects.all().delete()
            Investment.objects.all().delete()
            ArsUsdConversion.objects.all().delete()
            Investor.objects.all().delete()

        stats = {
            "investors": 0,
            "investments": 0,
            "expenses": 0,
            "commitments": 0,
            "works": 0,
            "participations": 0,
            "reinvestments": 0,
            "skipped_investments": 0,
            "skipped_works": 0,
            "skipped_participations": 0,
        }

        # ------------------------------------------------------------------
        # Investors bootstrap from known names
        # ------------------------------------------------------------------
        investors_by_name: dict[str, Investor] = {}

        def get_or_create_investor(name_raw: str) -> Investor:
            name = s(name_raw)
            if not name:
                name = "Sin asignar"
            if name not in investors_by_name:
                obj, _ = Investor.objects.get_or_create(name=name, defaults={"active": True})
                investors_by_name[name] = obj
            return investors_by_name[name]

        # ------------------------------------------------------------------
        # Commitments (hoja Pagos comprometidos)
        # ------------------------------------------------------------------
        ws_c = wb["Pagos comprometidos"]
        for r in range(2, ws_c.max_row + 1):
            due_date = dt(ws_c.cell(r, 1).value)
            concept = s(ws_c.cell(r, 2).value)
            amount = d(ws_c.cell(r, 3).value)
            currency_raw = s(ws_c.cell(r, 4).value).upper()
            usd_est = d(ws_c.cell(r, 5).value)
            status_raw = s(ws_c.cell(r, 6).value).lower()
            comment = s(ws_c.cell(r, 7).value)

            if not (due_date and concept and amount and currency_raw):
                continue

            currency = Currency.USD if currency_raw == "USD" else Currency.ARS

            if status_raw.startswith("pag"):
                status = Commitment.Status.PAID
            elif status_raw.startswith("pen"):
                status = Commitment.Status.PENDING
            else:
                status = Commitment.Status.PENDING

            Commitment.objects.create(
                due_date=due_date,
                concept=concept,
                amount=amount,
                currency=currency,
                estimated_usd=usd_est if usd_est is not None else Decimal("0"),
                status=status,
                comments=comment,
            )
            stats["commitments"] += 1

        # ------------------------------------------------------------------
        # Expenses (hoja Gastos realizados)
        # ------------------------------------------------------------------
        ws_g = wb["Gastos realizados"]
        for r in range(2, ws_g.max_row + 1):
            paid_date = dt(ws_g.cell(r, 1).value)
            concept = s(ws_g.cell(r, 2).value)
            fx = d(ws_g.cell(r, 3).value)
            amount = d(ws_g.cell(r, 4).value)
            currency_raw = s(ws_g.cell(r, 5).value).upper()
            usd = d(ws_g.cell(r, 6).value)
            payer_name = s(ws_g.cell(r, 7).value)
            comment = s(ws_g.cell(r, 8).value)

            if not (paid_date and concept and amount and currency_raw):
                continue

            payer = get_or_create_investor(payer_name)
            currency = Currency.USD if currency_raw == "USD" else Currency.ARS

            Expense.objects.create(
                date=paid_date,
                concept=concept,
                amount=amount,
                currency=currency,
                fx_ars_usd=fx if fx and fx > 0 else Decimal("1") if currency == Currency.USD else Decimal("0.0001"),
                amount_usd=usd if usd is not None else amount,
                comments=comment,
                payer=payer,
            )
            stats["expenses"] += 1

        # ------------------------------------------------------------------
        # Build helper to infer investor from commitment comment (optional)
        # ------------------------------------------------------------------
        commitment_hint: dict[tuple[date, str], Investor] = {}
        for c in Commitment.objects.all():
            key = (c.due_date, c.concept.lower())
            comment = c.comments.lower()
            matched = [inv for inv in Investor.objects.all() if inv.name.lower() in comment]
            if len(matched) == 1:
                commitment_hint[key] = matched[0]

        fallback_investor = get_or_create_investor("Sin asignar")

        # ------------------------------------------------------------------
        # Investments (hoja Inversión)
        # Header is row 2, data starts row 3
        # ------------------------------------------------------------------
        ws_i = wb["Inversión"]
        for r in range(3, ws_i.max_row + 1):
            inv_date = dt(ws_i.cell(r, 1).value)
            concept = s(ws_i.cell(r, 2).value)
            category = s(ws_i.cell(r, 3).value)
            amount = d(ws_i.cell(r, 4).value)
            currency_raw = s(ws_i.cell(r, 5).value).upper()
            fx = d(ws_i.cell(r, 6).value)
            usd = d(ws_i.cell(r, 7).value)
            payment_method = s(ws_i.cell(r, 10).value)
            term_detail = s(ws_i.cell(r, 11).value)
            status_raw = s(ws_i.cell(r, 12).value).lower()

            if not (inv_date and concept and amount and currency_raw):
                continue

            currency = Currency.USD if currency_raw == "USD" else Currency.ARS
            if status_raw.startswith("pag"):
                status = Investment.Status.PAID
            else:
                status = Investment.Status.PENDING

            payer = commitment_hint.get((inv_date, concept.lower()), fallback_investor)

            if fx is None:
                fx = Decimal("1") if currency == Currency.USD else Decimal("0.0001")
            if usd is None:
                usd = amount if currency == Currency.USD else Decimal("0")

            try:
                Investment.objects.create(
                    date=inv_date,
                    concept=concept,
                    category=category,
                    amount=amount,
                    currency=currency,
                    fx_ars_usd=fx,
                    amount_usd=usd,
                    payment_method=payment_method,
                    term_detail=term_detail,
                    status=status,
                    payer=payer,
                    comments="Importado desde Excel",
                )
                stats["investments"] += 1
            except Exception:
                stats["skipped_investments"] += 1

        # ------------------------------------------------------------------
        # Works (hoja Trabajos) + participations from cols 11-17
        # ------------------------------------------------------------------
        ws_t = wb["Trabajos"]
        works_by_sheet: dict[str, Work] = {}
        used_work_ids: set[str] = set()

        for r in range(2, ws_t.max_row + 1):
            work_date = dt(ws_t.cell(r, 1).value)
            sheet_name = s(ws_t.cell(r, 2).value)
            if not (work_date and sheet_name):
                continue

            m = re.search(r"T\\s*-?\\s*(\\d+)", sheet_name.upper())
            if m:
                work_id = f"T{m.group(1)}"
            else:
                work_id = f"W{r}"
            work_id = work_id[:20]
            suffix = 1
            base_id = work_id
            while work_id in used_work_ids:
                work_id = f"{base_id[:17]}-{suffix}"[:20]
                suffix += 1
            used_work_ids.add(work_id)

            status_raw = s(ws_t.cell(r, 8).value).lower()
            if status_raw.startswith("cob"):
                status = Work.Status.COLLECTED
            elif status_raw.startswith("fac"):
                status = Work.Status.BILLED
            elif status_raw.startswith("pen"):
                status = Work.Status.PENDING
            else:
                status = Work.Status.PENDING

            work = Work.objects.create(
                date=work_date,
                work_id=work_id,
                hectares=d(ws_t.cell(r, 3).value),
                work_type=s(ws_t.cell(r, 4).value) or "Sin definir",
                usd_per_hectare=d(ws_t.cell(r, 5).value),
                total_ars=d(ws_t.cell(r, 6).value) or Decimal("0"),
                total_usd=d(ws_t.cell(r, 7).value) or Decimal("0"),
                status=status,
                comments=s(ws_t.cell(r, 9).value),
            )
            works_by_sheet[sheet_name] = work
            stats["works"] += 1

        for r in range(2, ws_t.max_row + 1):
            p_date = dt(ws_t.cell(r, 11).value)
            p_sheet = s(ws_t.cell(r, 12).value)
            p_investor_name = s(ws_t.cell(r, 13).value)
            p_pct = d(ws_t.cell(r, 14).value)
            p_reason = s(ws_t.cell(r, 15).value)
            p_amount = d(ws_t.cell(r, 16).value)
            p_destination_raw = s(ws_t.cell(r, 17).value).lower()

            if not (p_date and p_sheet and p_investor_name and p_amount):
                continue

            work = works_by_sheet.get(p_sheet)
            if work is None:
                stats["skipped_participations"] += 1
                continue

            investor = get_or_create_investor(p_investor_name)

            destination = (
                WorkParticipation.Destination.REINVESTMENT
                if "reinv" in p_destination_raw
                else WorkParticipation.Destination.PAYMENT
            )

            percentage = (p_pct * Decimal("100")) if p_pct is not None and p_pct <= 1 else (p_pct or Decimal("0"))

            WorkParticipation.objects.create(
                date=p_date,
                work=work,
                investor=investor,
                percentage=percentage,
                reason=p_reason,
                amount_usd=p_amount,
                destination=destination,
            )
            stats["participations"] += 1

            if destination == WorkParticipation.Destination.REINVESTMENT:
                Reinvestment.objects.create(
                    date=p_date,
                    investor=investor,
                    amount_usd=p_amount,
                    comments=f"Importado desde participación de trabajo: {work.work_id}",
                )
                stats["reinvestments"] += 1

        # Ensure at least one conversion exists for fallback FX service
        if not ArsUsdConversion.objects.exists():
            ArsUsdConversion.objects.create(
                date=date.today(),
                ars_amount=Decimal("1"),
                fx_ars_usd=Decimal("1200"),
                usd_obtained=Decimal("0.0008"),
                bank_detail="Fallback inicial",
                comments="Generado automáticamente",
            )

        stats["investors"] = Investor.objects.count()

        self.stdout.write(self.style.SUCCESS("Importación inicial completada."))
        for k, v in stats.items():
            self.stdout.write(f"- {k}: {v}")
